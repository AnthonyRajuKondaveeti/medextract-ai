"""
excel_writer.py

Builds the final Excel workbook from extracted patient records.

Single sheet — "Results"
    S.No prepended automatically.
    One row per patient, columns in the agreed order.
    _Flag columns are internal only — not written as separate columns.
    Flags are embedded in the value cell:
        HIGH → "12.6 (H)"
        LOW  → "12.6 (L)"
    No colours. Frozen header row. Auto-filter. Auto-sized columns.
"""

import io
import logging
import re
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from validator import MASTER_COLUMNS, COLUMN_DISPLAY_NAMES, FLAG_FIELDS

logger = logging.getLogger(__name__)

_FONT_NAME   = "Calibri"
_FONT_HEADER = Font(name=_FONT_NAME, bold=True, size=10)
_FONT_BODY   = Font(name=_FONT_NAME, size=10)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
_ALIGN_WRAP   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# Graph/image fields — display attachment status instead of content
_GRAPH_FIELDS = {"XRAY", "PFT", "AUDIOMETRY"}


# ── Data Normalization ─────────────────────────────────────────────────────────

def _normalize_spelling(text: str) -> str:
    """
    Normalize spelling variations (case, punctuation, extra spaces).
    Example: "NORMAL STUDY." → "Normal Study"
    """
    if not text:
        return text
    
    # Remove trailing punctuation
    text = text.rstrip('.,;:!?')
    
    # Normalize whitespace
    text = ' '.join(text.split())
    
    # Title case for common medical terms
    text_upper = text.upper()
    
    # Keep acronyms uppercase, title case for words
    if text_upper in ['NORMAL', 'ABNORMAL', 'NAD', 'WNL', 'NIL', 'ABSENT', 'PRESENT']:
        return text_upper
    
    # Title case for mixed text
    return text.title()


def _normalize_multi_value(value: str | None) -> str | None:
    """
    Normalize fields with pipe-separated multiple values.
    Returns first unique non-empty value with spelling normalization.
    Example: "Normal | NORMAL | NORMAL STUDY." → "NORMAL"
    """
    if not value or not isinstance(value, str):
        return value
    
    if '|' not in value:
        return _normalize_spelling(value.strip()) if value.strip() else None
    
    parts = [p.strip() for p in value.split('|')]
    
    # Normalize spelling for each part
    normalized_parts = [_normalize_spelling(p) for p in parts if p]
    
    # Remove duplicates (case-insensitive) while preserving order
    seen = set()
    unique = []
    for part in normalized_parts:
        part_lower = part.lower()
        if part and part_lower not in seen:
            seen.add(part_lower)
            unique.append(part)
    
    return unique[0] if unique else None


def _normalize_date(date_str: str | None) -> str | None:
    """
    Parse multiple date formats and return standardized DD-MM-YYYY.
    Handles pipe-separated dates and various formats.
    """
    if not date_str or not isinstance(date_str, str):
        return date_str
    
    dates = [d.strip() for d in date_str.split('|')]
    
    formats = [
        '%d-%m-%y', '%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d',
        '%d-%b-%y %I.%M %p', '%d-%b-%Y %I.%M %p',
        '%d-%m-%Y %I:%M %p', '%d/%m/%Y %I:%M %p',
    ]
    
    parsed_dates = []
    for date_item in dates:
        # Handle "11th December 2025" format
        month_name_match = re.search(r'(\d+)(?:st|nd|rd|th)?\s+([A-Za-z]+)\s+(\d{4})', date_item)
        if month_name_match:
            try:
                day, month_name, year = month_name_match.groups()
                parsed = datetime.strptime(f"{day} {month_name} {year}", "%d %B %Y")
                parsed_dates.append(parsed)
                continue
            except:
                pass
        
        # Try standard formats
        for fmt in formats:
            try:
                parsed = datetime.strptime(date_item, fmt)
                parsed_dates.append(parsed)
                break
            except:
                continue
    
    if parsed_dates:
        # Return most recent date in DD-MM-YYYY format (Indian standard)
        most_recent = max(parsed_dates)
        return most_recent.strftime('%d-%m-%Y')
    
    # If parsing fails, return first value
    return dates[0] if dates else None


def _normalize_record(record: dict) -> dict:
    """
    Normalize a patient record by cleaning multi-value fields.
    Applied before writing to Excel.
    """
    normalized = record.copy()
    
    # Fields that commonly have multiple values
    multi_value_fields = ['Lab_Name', 'Mobile', 'UHIDNo', 'EmpCode']
    
    # Medical report fields that need spelling normalization
    medical_fields = ['XRAY', 'PFT', 'AUDIOMETRY', 'Remarks', 'Suggestion']
    
    for field in multi_value_fields:
        if field in normalized:
            normalized[field] = _normalize_multi_value(normalized[field])
    
    for field in medical_fields:
        if field in normalized:
            normalized[field] = _normalize_multi_value(normalized[field])
    
    # Normalize date field
    if 'Report_Date' in normalized:
        normalized['Report_Date'] = _normalize_date(normalized['Report_Date'])
    
    return normalized

# ── Column configuration ───────────────────────────────────────────────────────

# All MASTER_COLUMNS except internal _Flag columns
_OUTPUT_COLUMNS: list[str] = [
    col for col in MASTER_COLUMNS
    if not col.endswith("_Flag")
]

# Map value column → its flag column where one exists
_FLAG_MAP: dict[str, str] = {
    col: f"{col}_Flag"
    for col in _OUTPUT_COLUMNS
    if f"{col}_Flag" in FLAG_FIELDS
}

# Long free-text columns — use wrap text
_WRAP_COLS = {
    "Remarks", "Suggestion",
    "Extraction_Note", "Data_Quality",
}

# Column widths — fixed for known columns, derived from header for the rest
_FIXED_WIDTHS: dict[str, int] = {
    "EmpCode":          12,
    "UHIDNo":           12,
    "PatientName":      22,
    "Age":               6,
    "Gender":            8,
    "Height":            8,
    "Weight":            8,
    "BMI":               7,
    "BP":               10,
    "Pulse":             7,
    "Mobile":           14,
    "Blood_Group":      10,
    "Rh_Type":          10,
    "XRAY":             15,  # Shows "Attached" / "Not Attached"
    "PFT":              15,  # Shows "Attached" / "Not Attached"
    "AUDIOMETRY":       15,  # Shows "Attached" / "Not Attached"
    "Remarks":          30,
    "Suggestion":       30,
    "Extraction_Note":  35,
    "Data_Quality":     35,
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _embed_flag(value, flag: Optional[str]) -> str:
    """
    Combine value and flag into a single display string.
        12.6 + HIGH → "12.6 (H)"
        12.6 + LOW  → "12.6 (L)"
        12.6 + None → "12.6"
        None + any  → ""
    """
    if value is None:
        return ""
    s = str(value)
    if flag == "HIGH":
        return f"{s} (H)"
    if flag == "LOW":
        return f"{s} (L)"
    return s


def _col_width(col_name: str, display_name: str) -> int:
    if col_name in _FIXED_WIDTHS:
        return _FIXED_WIDTHS[col_name]
    return max(len(display_name) + 2, 10)


# ── Sheet builder ──────────────────────────────────────────────────────────────

def _write_results_sheet(sheet, records: list[dict]) -> None:
    # Header row: S.No + display name per output column
    headers = ["S.No"] + [
        COLUMN_DISPLAY_NAMES.get(col, col) for col in _OUTPUT_COLUMNS
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font      = _FONT_HEADER
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 30
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Data rows
    for row_idx, raw_record in enumerate(records, start=2):
        # Normalize record (clean multi-value fields)
        record = _normalize_record(raw_record)
        
        # S.No
        cell = sheet.cell(row=row_idx, column=1, value=row_idx - 1)
        cell.font      = _FONT_BODY
        cell.alignment = _ALIGN_CENTER

        for col_idx, col_name in enumerate(_OUTPUT_COLUMNS, start=2):
            value    = record.get(col_name)
            flag_col = _FLAG_MAP.get(col_name)
            flag     = record.get(flag_col) if flag_col else None

            # Graph fields: show attachment status instead of content
            if col_name in _GRAPH_FIELDS:
                display = "Attached" if (value and str(value).strip()) else "Not Attached"
            elif flag_col:
                display = _embed_flag(value, flag) or None
            else:
                display = None if value is None else str(value)

            cell = sheet.cell(row=row_idx, column=col_idx, value=display)
            cell.font      = _FONT_BODY
            cell.alignment = _ALIGN_WRAP if col_name in _WRAP_COLS else _ALIGN_LEFT

    # Column widths
    sheet.column_dimensions["A"].width = 6
    for col_idx, col_name in enumerate(_OUTPUT_COLUMNS, start=2):
        display_name = COLUMN_DISPLAY_NAMES.get(col_name, col_name)
        sheet.column_dimensions[get_column_letter(col_idx)].width = (
            _col_width(col_name, display_name)
        )


# ── Public API ─────────────────────────────────────────────────────────────────

def build_excel(
    records: list[dict],
    summaries: list[dict],
) -> bytes:
    """
    Build the Excel workbook — single Results sheet.

    Args:
        records:   Validated patient dicts, one per PDF.
        summaries: Pipeline stats (stored in DB, not written to Excel).

    Returns:
        Raw .xlsx bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    _write_results_sheet(ws, records)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info(f"Excel built: {len(records)} patient row(s).")
    return buffer.read()


def get_output_filename() -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"medical_reports_{timestamp}.xlsx"
