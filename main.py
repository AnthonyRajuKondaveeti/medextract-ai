"""
main.py

FastAPI entry point for MedExtract AI.

Auth:
    Two methods accepted on all data endpoints:
      1. Bearer session token  (post /login, then pass Authorization: Bearer <token>)
      2. X-API-Key header      (legacy / programmatic access)

Sessions are stored in PostgreSQL — they survive container restarts and
rolling redeploys.  Admin credentials are loaded from environment variables;
the password is verified against a bcrypt hash.

Production process management:
    CMD in Dockerfile uses Gunicorn + UvicornWorker.
    For local dev only, python main.py still works.
"""

import asyncio
import logging
import os
import sys
import time
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional

import bcrypt

# Ensure src/ is on the path so all module imports resolve
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "src"))

from fastapi import BackgroundTasks, Depends, FastAPI, File, HTTPException, Request, Security, UploadFile, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, Response
from fastapi.security.api_key import APIKeyHeader
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from slowapi import Limiter
from slowapi.errors import RateLimitExceeded
from slowapi.util import get_remote_address
import secrets

from _db import (
    db_create_session,
    db_validate_session,
    db_delete_session,
    db_cleanup_sessions,
    db_get_all_completed_jobs,
    db_get_excel,
    close_db,
)
from batch_processor import (
    STATUS_COMPLETE,
    create_job,
    get_job,
    get_job_status_payload,
    init_db,
    run_batch,
)
from config import API_KEY, OUTPUT_DIR, UPLOAD_DIR


# ── Logging ────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)


# ── Session Management ─────────────────────────────────────────────────────────
# Sessions are persisted in PostgreSQL so they survive container restarts
# and rolling redeploys.  Expired rows are pruned hourly by _cleanup_task().

SESSION_TIMEOUT_MINUTES = 480  # 8 hours

# Admin credentials from env vars — never hardcoded.
# Generate hash: python -c "import bcrypt; print(bcrypt.hashpw(b'pw', bcrypt.gensalt()).decode())"
# Then set ADMIN_PASSWORD_HASH=<hash> in .env.docker / .env.prod
ADMIN_USERNAME: str = os.getenv("ADMIN_USERNAME", "admin")
_ADMIN_PASSWORD_HASH: str = os.getenv("ADMIN_PASSWORD_HASH", "")


def _check_password(password: str) -> bool:
    """Verify password against bcrypt hash stored in ADMIN_PASSWORD_HASH env var.
    Falls back to credential-free login only when ENV=dev and no hash is set.
    """
    if _ADMIN_PASSWORD_HASH:
        try:
            return bcrypt.checkpw(password.encode(), _ADMIN_PASSWORD_HASH.encode())
        except Exception:
            return False
    # No hash configured — allow login only in dev mode (no password required)
    logger.warning(
        "ADMIN_PASSWORD_HASH is not set. "
        "Login is allowed without a password only in ENV=dev. "
        "Set ADMIN_PASSWORD_HASH in production."
    )
    return os.getenv("ENV", "dev") == "dev"


class LoginRequest(BaseModel):
    username: str
    password: str


def create_session_token() -> str:
    """Generate a cryptographically secure random session token."""
    return secrets.token_urlsafe(32)


def validate_session_token(token: Optional[str]) -> bool:
    """Check if token exists in PostgreSQL and has not expired."""
    if not token:
        return False
    try:
        return db_validate_session(token)
    except Exception as exc:
        logger.warning(f"Session validation DB error: {exc}")
        return False


# ── Authentication ─────────────────────────────────────────────────────────────

_api_key_header = APIKeyHeader(name="X-API-Key", auto_error=False)


async def require_api_key(key: str = Security(_api_key_header)) -> None:
    """FastAPI dependency — raises 403 if API key is missing or wrong."""
    if not key or key != API_KEY:
        raise HTTPException(
            status_code=403,
            detail="Invalid or missing API key. Provide header: X-API-Key: <key>",
        )


async def require_auth(
    authorization: Optional[str] = Header(None),
    api_key: Optional[str] = Security(_api_key_header),
) -> None:
    """Accept either a valid Bearer session token or a valid API key."""
    if authorization and authorization.startswith("Bearer "):
        token = authorization.replace("Bearer ", "", 1)
        if validate_session_token(token):
            return
    if api_key and api_key == API_KEY:
        return
    raise HTTPException(
        status_code=401,
        detail="Authentication required. Please login or provide a valid API key.",
    )


_auth = Depends(require_auth)


# ── Rate Limiter ───────────────────────────────────────────────────────────────

_limiter = Limiter(key_func=get_remote_address)


def _rate_limit_handler(request: Request, exc: RateLimitExceeded) -> JSONResponse:
    return JSONResponse(
        status_code=429,
        content={"detail": f"Rate limit exceeded. {exc.detail}"},
    )


# ── Upload limits ──────────────────────────────────────────────────────────────

_MAX_FILE_BYTES = int(os.getenv("MAX_FILE_SIZE_MB", "50")) * 1024 * 1024


# ── Background cleanup ─────────────────────────────────────────────────────────

async def _cleanup_task() -> None:
    """Hourly background task: purge expired DB sessions and stale upload files."""
    while True:
        await asyncio.sleep(3600)

        # Clean expired sessions
        try:
            removed = await asyncio.to_thread(db_cleanup_sessions)
            if removed:
                logger.info(f"Session cleanup: removed {removed} expired session(s).")
        except Exception as exc:
            logger.warning(f"Session cleanup error: {exc}")

        # Clean upload files older than 24 hours
        try:
            cutoff = time.time() - 24 * 3600
            count = 0
            for f in Path(UPLOAD_DIR).iterdir():
                if f.is_file() and f.stat().st_mtime < cutoff:
                    f.unlink(missing_ok=True)
                    count += 1
            if count:
                logger.info(f"Upload cleanup: removed {count} stale file(s).")
        except Exception as exc:
            logger.warning(f"Upload cleanup error: {exc}")


# ── App lifecycle ──────────────────────────────────────────────────────────────

@asynccontextmanager
async def lifespan(app: FastAPI):
    # ── Production startup guards ──────────────────────────────────────────────
    # These checks only run when ENV=production and will abort startup on
    # misconfiguration rather than silently running in an insecure state.
    _env = os.getenv("ENV", "dev")
    if _env == "production":
        if not _ADMIN_PASSWORD_HASH:
            raise RuntimeError(
                "FATAL: ADMIN_PASSWORD_HASH is not set but ENV=production. "
                "Generate a bcrypt hash with:\n"
                "  python -c \"import bcrypt; "
                "print(bcrypt.hashpw(b'YourPassword', bcrypt.gensalt()).decode())\"\n"
                "Then set ADMIN_PASSWORD_HASH=<hash> in your .env.prod file."
            )
        _raw_origins = os.getenv("ALLOWED_ORIGINS", "*")
        if _raw_origins.strip() == "*":
            raise RuntimeError(
                "FATAL: ALLOWED_ORIGINS=* is not allowed in production (ENV=production). "
                "Set ALLOWED_ORIGINS to your actual domain(s) in .env.prod, e.g.:\n"
                "  ALLOWED_ORIGINS=https://your-domain.com"
            )
        logger.info("Production startup checks passed.")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    init_db()   # creates/verifies schema including sessions table (idempotent)
    asyncio.create_task(_cleanup_task())
    logger.info("Medical Extractor API started.")
    yield
    # Graceful shutdown: close DB connection pool
    close_db()
    logger.info("Medical Extractor API shutting down.")


# ── App setup ─────────────────────────────────────────────────────────────────

app = FastAPI(
    title="Medical Report Extraction Platform",
    description="Extracts structured data from Indian lab PDF reports using OpenAI.",
    version="2.0.0",
    lifespan=lifespan,
)

app.state.limiter = _limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_handler)

# CORS — restrict origins in production via ALLOWED_ORIGINS env var.
# Example: ALLOWED_ORIGINS=https://yourapp.com,https://admin.yourapp.com
_ALLOWED_ORIGINS = [o.strip() for o in os.getenv("ALLOWED_ORIGINS", "*").split(",")]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST"],
    allow_headers=["X-API-Key", "Authorization", "Content-Type"],
)

# Serve static frontend — no auth required so browser can load the UI
app.mount(
    "/static",
    StaticFiles(directory=os.path.join(os.path.dirname(__file__), "static")),
    name="static",
)


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.get("/", include_in_schema=False)
async def serve_index():
    """Serve the frontend SPA — no auth required."""
    index_path = os.path.join(os.path.dirname(__file__), "static", "index.html")
    if not os.path.exists(index_path):
        raise HTTPException(status_code=404, detail="Frontend not found.")
    return FileResponse(index_path)


@app.get("/health", include_in_schema=False)
async def health_check():
    """Health probe — no auth required. Used by load balancers and Docker HEALTHCHECK."""
    return {"status": "ok"}


@app.post("/login")
async def login(credentials: LoginRequest):
    """
    Authenticate with username/password and receive a session token.
    Returns: { "token": "<session_token>", "expires_in": 28800 }
    """
    if credentials.username != ADMIN_USERNAME or not _check_password(credentials.password):
        raise HTTPException(status_code=401, detail="Invalid username or password.")

    token = create_session_token()
    expiry = datetime.now(tz=timezone.utc) + timedelta(minutes=SESSION_TIMEOUT_MINUTES)
    await asyncio.to_thread(db_create_session, token, expiry)

    logger.info(f"User '{credentials.username}' logged in.")
    return {
        "token": token,
        "expires_in": SESSION_TIMEOUT_MINUTES * 60,
        "message": "Login successful",
    }


@app.post("/logout")
async def logout(authorization: Optional[str] = Header(None)):
    """Invalidate current session token."""
    if authorization and authorization.startswith("Bearer "):
        token = authorization.replace("Bearer ", "", 1)
        await asyncio.to_thread(db_delete_session, token)
        return {"message": "Logged out successfully"}
    return {"message": "No active session to logout"}


@app.post("/upload", dependencies=[_auth])
@_limiter.limit("20/minute")
async def upload_pdfs(
    request: Request,
    background_tasks: BackgroundTasks,
    files: list[UploadFile] = File(...),
):
    """
    Accept one or more PDF files and start processing.
    Requires: Authorization: Bearer <token>  OR  X-API-Key: <key>
    Rate limited to 20 requests / IP / minute.
    Returns: { "job_id": "<uuid>" }
    """
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded.")

    file_payloads: list[tuple[str, bytes]] = []
    for upload in files:
        if not upload.filename:
            raise HTTPException(status_code=400, detail="File missing filename.")

        content = await upload.read()

        if not content:
            raise HTTPException(
                status_code=400,
                detail=f"File '{upload.filename}' is empty.",
            )

        # File size guard
        if len(content) > _MAX_FILE_BYTES:
            raise HTTPException(
                status_code=413,
                detail=(
                    f"'{upload.filename}' is too large "
                    f"({len(content) // (1024*1024)} MB). "
                    f"Maximum allowed: {_MAX_FILE_BYTES // (1024*1024)} MB."
                ),
            )

        # PDF magic bytes validation — rejects non-PDF files early
        if not content.startswith(b"%PDF"):
            raise HTTPException(
                status_code=415,
                detail=f"'{upload.filename}' is not a valid PDF file.",
            )

        file_payloads.append((upload.filename, content))

    filenames = [fp[0] for fp in file_payloads]
    job_id = create_job(filenames)

    background_tasks.add_task(_run_batch_background, job_id, file_payloads)

    logger.info(f"Job {job_id} queued with {len(file_payloads)} file(s).")
    return JSONResponse(content={"job_id": job_id}, status_code=202)


async def _run_batch_background(
    job_id: str,
    file_payloads: list[tuple[str, bytes]],
) -> None:
    """Background coroutine that drives the batch processor."""
    try:
        await run_batch(job_id, file_payloads)
    except Exception as exc:
        logger.error(f"Job {job_id} background task crashed: {exc}", exc_info=True)


@app.get("/status/{job_id}", dependencies=[_auth])
async def get_status(job_id: str):
    """Poll job processing status."""
    payload = get_job_status_payload(job_id)
    if payload is None:
        raise HTTPException(status_code=404, detail=f"Job '{job_id}' not found.")
    return JSONResponse(content=payload)


@app.get("/download/all", dependencies=[_auth])
async def download_all_batches():
    """
    Download a combined Excel report containing all patient records from all completed jobs.
    Combines all batches into a single Excel file.
    """
    from openpyxl import load_workbook
    from excel_writer import build_excel, get_output_filename
    import io

    # Get all completed jobs
    completed_jobs = await asyncio.to_thread(db_get_all_completed_jobs)
    
    if not completed_jobs:
        raise HTTPException(
            status_code=404,
            detail="No completed jobs found. Process some files first.",
        )

    # Collect all patient records from all jobs
    all_records = []
    jobs_processed = 0

    for job_info in completed_jobs:
        job_id = job_info["job_id"]
        excel_bytes = await asyncio.to_thread(db_get_excel, job_id)
        
        if not excel_bytes:
            logger.warning(f"Job {job_id} has no Excel data, skipping.")
            continue

        try:
            # Parse the Excel file to extract patient records
            wb = load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
            ws = wb["Results"]  # The results sheet

            # Read header row
            headers = [cell.value for cell in ws[1]]
            
            # Skip first column (S.No) and read data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Skip empty rows
                    continue
                    
                # Build record dict (skip S.No column)
                record = {}
                for idx, header in enumerate(headers[1:], start=1):
                    value = row[idx] if idx < len(row) else None
                    
                    # Handle embedded flags like "12.6 (H)" or "12.6 (L)"
                    if value and isinstance(value, str):
                        # Remove flag annotations for storage
                        value = value.replace(" (H)", "").replace(" (L)", "")
                        # Handle "Not Attached" / "Attached" for graph fields
                        if value in ["Not Attached", "Attached"]:
                            value = None if value == "Not Attached" else "PRESENT"
                    
                    # Map display name back to internal field name
                    # This is a reverse lookup of COLUMN_DISPLAY_NAMES
                    from validator import COLUMN_DISPLAY_NAMES
                    field_name = None
                    for k, v in COLUMN_DISPLAY_NAMES.items():
                        if v == header:
                            field_name = k
                            break
                    
                    if field_name:
                        record[field_name] = value
                    else:
                        record[header] = value
                
                all_records.append(record)
            
            wb.close()
            jobs_processed += 1
            
        except Exception as exc:
            logger.error(f"Error parsing Excel for job {job_id}: {exc}")
            continue

    if not all_records:
        raise HTTPException(
            status_code=404,
            detail="No patient records found in completed jobs.",
        )

    # Build combined Excel file
    try:
        combined_excel = await asyncio.to_thread(
            build_excel, all_records, []  # Empty summaries for combined report
        )
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"all_batches_combined_{timestamp}.xlsx"
        
        logger.info(
            f"Combined Excel generated: {len(all_records)} records from {jobs_processed} job(s)."
        )
        
        return Response(
            content=combined_excel,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Length": str(len(combined_excel)),
            },
        )
    except Exception as exc:
        logger.error(f"Error building combined Excel: {exc}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Failed to build combined Excel: {str(exc)}",
        )


@app.get("/download/{job_id}", dependencies=[_auth])
async def download_excel(job_id: str):
    """
    Download the generated Excel report for a completed job.
    Only available when job status is 'complete'.
    """
    job = get_job(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail=f"Job '{job_id}' not found.")

    if job.status != STATUS_COMPLETE:
        raise HTTPException(
            status_code=409,
            detail=f"Job is not complete yet. Current status: '{job.status}'.",
        )

    if not job.excel_bytes:
        raise HTTPException(
            status_code=500,
            detail="Excel file could not be generated for this job.",
        )

    filename = job.excel_filename or "medical_reports.xlsx"

    return Response(
        content=job.excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(job.excel_bytes)),
        },
    )


# ── Entry point ────────────────────────────────────────────────────────────────
# In production, Gunicorn is used (see Dockerfile CMD).
# This block is for local dev only: python main.py

if __name__ == "__main__":
    import uvicorn

    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=8000,
        reload=False,
        log_level="info",
    )
